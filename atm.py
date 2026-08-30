import os
from openpyxl import Workbook, load_workbook

file = "atm_accounts.xlsx"

if os.path.exists(file):
    wb = load_workbook(file)
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.activehttps://github.com/notifications
    sheet.append(["Name", "Address", "Contact", "Bank Balance"])
    wb.save(file)

current_user_row = None

def create_account():
    print("\n===== CREATE ACCOUNT =====")
    name = input("Enter your name: ")
    address = input("Enter your address: ")
    contact = input("Enter your contact number: ")
    
    try:
        balance = float(input("Enter your initial bank balance: "))
    except ValueError:
        print("Invalid amount. Account creation failed.")
        return

    sheet.append([name, address, contact, balance])
    wb.save(file)
    print("\nYour account has been created successfully!")

def login():
    global current_user_row
    print("\n===== LOGIN =====")
    name = input("Enter your registered name to log in: ").strip()
    
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == name:
            current_user_row = row
            print(f"Welcome back, {name}!")
            return True
            
    print("Account not found! Please create an account first.")
    return False

def deposit():
    if current_user_row is None:
        print("Please log in first!")
        return
        
    print("\n===== DEPOSIT =====")
    try:
        money = float(input("Enter the money to deposit: "))
        if money <= 0:
            print("Amount must be greater than zero.")
            return
    except ValueError:
        print("Invalid amount.")
        return

    current_balance = float(sheet.cell(current_user_row, 4).value or 0)
    new_balance = current_balance + money
    
    sheet.cell(current_user_row, 4).value = new_balance
    wb.save(file)
    print("Your money has been deposited successfully!")
    print("Your updated bank balance is:", new_balance)

def withdraw():
    if current_user_row is None:
        print("Please log in first!")
        return
        
    print("\n===== WITHDRAW =====")
    try:
        money = float(input("Enter amount to withdraw: "))
        if money <= 0:
            print("Amount must be greater than zero.")
            return
    except ValueError:
        print("Invalid amount.")
        return

    current_balance = float(sheet.cell(current_user_row, 4).value or 0)
    if money > current_balance:
        print("Insufficient balance!")
        return
        
    new_balance = current_balance - money
    sheet.cell(current_user_row, 4).value = new_balance
    wb.save(file)
    print("Your money has been withdrawn successfully!")
    print("Your updated bank balance is:", new_balance)

def check_balance():
    if current_user_row is None:
        print("Please log in first!")
        return
    balance = sheet.cell(current_user_row, 4).value
    print("\nYour current bank balance is:", balance)

while True:
    print("\n===== ATM MACHINE =====")
    print("1. Create Account")
    print("2. Log In to Existing Account")
    print("3. Deposit Money")
    print("4. Withdraw Money")
    print("5. Check Balance")
    print("6. Exit")
    
    try:
        choice = int(input("Enter your choice: "))
    except ValueError:
        print("Please enter a valid number.")
        continue
        
    if choice == 1:
        create_account()
    elif choice == 2:
        login()
    elif choice == 3:
        deposit()
    elif choice == 4:
        withdraw()
    elif choice == 5:
        check_balance()
    elif choice == 6:
        print("Thank you for using the ATM!")
        break
    else:
        print("Invalid choice!")
